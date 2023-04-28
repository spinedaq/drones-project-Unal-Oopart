#!/usr/bin/python3
import tkinter
import tkinter.messagebox
from tkinter import filedialog
from tkinter import ttk
import customtkinter
import os
import time
import multiprocessing
from PIL import Image
import argparse
import platform
import sys
import math
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import datetime
from datetime import time, timedelta
import torch
import pandas as pd


#Yolo
FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from models.common import DetectMultiBackend
from utils.dataloaders import IMG_FORMATS, VID_FORMATS, LoadImages, LoadScreenshots, LoadStreams
from utils.general import (LOGGER, Profile, check_file, check_img_size, check_imshow, check_requirements, colorstr, cv2,
                           increment_path, non_max_suppression, print_args, scale_boxes, strip_optimizer, xyxy2xywh)
from utils.plots import Annotator, colors, save_one_box
from utils.torch_utils import select_device, smart_inference_mode

#Tkinter
customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Programa de detección")
        self.geometry(f"{1285}x{650}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)
        customtkinter.set_widget_scaling(0.8)

        # load images side bar
        self.image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "icono_logotipo.png")), size=(150, 150))
        self.cover_home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "portada_inicio_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "portada_inicio_dark.png")), size=(910, 650))
        self.icon_home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "icono_inicio_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "icono_inicio_dark.png")), size=(45, 45))
        self.icon_image_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "icono_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "icono_imagen_dark.png")), size=(45, 45))
        self.icon_video_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "icono_video_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "icono_video_dark.png")), size=(45, 45))
        self.icon_streaming_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "icono_streaming_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "icono_streaming_dark.png")), size=(45, 45))
                                             

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(5, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="", image=self.logo_image, compound="top", font=("Microsoft GothicNeo", 20))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Seleccionar apariencia:", anchor="w")
        self.appearance_mode_label.grid(row=6, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Claro", "Oscuro", "Sistema"],command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=7, column=0, padx=20, pady=(10, 10))
        
        self.home_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, border_width=0, height=40, border_spacing=10, text="Inicio",
                                                  fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                 image=self.icon_home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, padx=50, pady=10)

        self.image_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, border_width=0, height=40, border_spacing=10, text="Imagen",
                                                  fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                 image=self.icon_image_image, anchor="w", command=self.image_button_event)
        self.image_button.grid(row=2, column=0, padx=50, pady=10)
        
        self.video_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Video",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.icon_video_image, anchor="w", command=self.video_button_event)
        self.video_button.grid(row=3, column=0, padx=50, pady=10)

        self.streaming_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="En vivo",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.icon_streaming_image, anchor="w", command=self.streaming_button_event)
        self.streaming_button.grid(row=4, column=0, padx=50, pady=10) 


        # create home frame
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)
        self.cover_image = customtkinter.CTkLabel(self.home_frame, text="", image=self.cover_home_image)
        self.cover_image.grid(row=0, column=0)
        

        # create image frame
        self.image_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.image_frame.grid_columnconfigure(5, weight=1)
        self.image_frame.grid_rowconfigure(5, weight=1)

        self.source_selector_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_recursos_imagen.png")), size=(335, 53))
        self.source_selector = customtkinter.CTkLabel(self.image_frame, text="", image=self.source_selector_image)
        self.source_selector.grid(row=0, column=0, columnspan=2, padx=(0,20), pady=(19,0),sticky = "we")

        self.icon_upload_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_dark.png")), size=(45, 45))
        self.image_source_selector_button_1 = customtkinter.CTkButton(self.image_frame, text="SELECCIONAR CARPETA \n DE RECURSOS", image=self.icon_upload_files_image, compound="bottom", command= self.image_source_selector_button_1_event, height = 250)
        self.image_source_selector_button_1.grid(row=1, column=0, pady=(20,0)) 

        self.image_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Elementos seleccionados")
        self.image_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.image_select_source_folder_files = "Sin seleccionar" #Default value
        self.image_source_scrollable_elements = customtkinter.CTkLabel(master = self.image_source_scrollable_frame, text = self.image_select_source_folder_files, anchor="center")
        self.image_source_scrollable_elements.grid(row=0, column=0, padx = (55,0))

       
        self.source_save_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_guardado_imagen.png")), size=(335, 53))
        self.source_save = customtkinter.CTkLabel(self.image_frame, text="", image=self.source_save_image)
        self.source_save.grid(row=2, column=0, columnspan=2, padx=(0,20), pady=(20,0),sticky = "we")

        self.icon_download_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_dark.png")), size=(45, 45))
        self.image_source_save_button_1 = customtkinter.CTkButton(self.image_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.image_source_save_button_1_event, height=165)
        self.image_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.image_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.image_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.image_select_folder = "Sin seleccionar" #Default value
        self.image_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.image_source_save_scrollable_frame, text = self.image_select_folder, anchor = "w", compound="left")
        self.image_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))


       
        self.visual_inference_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "visualizacion_inferencia_imagen.png")), size=(442, 53))
        self.visual_inference = customtkinter.CTkLabel(self.image_frame, text="", image=self.visual_inference_image)
        self.visual_inference.grid(row=0, column=2, padx=(0,0), pady=(19,0),sticky = "we")
        
        self.image_inference_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "inferencia_image_light.png")), dark_image=Image.open(os.path.join(self.image_path, "inferencia_image_dark.png")),size=(439, 247))
        self.inference_image_image = customtkinter.CTkLabel(self.image_frame, text="", image=self.image_inference_image, compound="center", fg_color=('#3b8ed0','#3b8ed0')
                                                                                              , corner_radius=8, height=262)
        self.inference_image_image.grid(row=1, column=2, padx=(0,0), pady=(20, 0))

        self.progressbar_inference_image = customtkinter.CTkLabel(self.image_frame, text="PROGRESO:", font=("Microsoft GothicNeo Bold", 12), anchor = "w")
        self.progressbar_inference_image.grid(row=2, column=2, padx=(0,325))

        self.progressbar_image_frame = customtkinter.CTkProgressBar(self.image_frame, height = 20)
        self.progressbar_image_frame.grid(row=2, column=2, padx=(0,40)) #padx=(10, 20), pady=(10, 10), 
        self.progressbar_image_frame.set(0.75)

        self.progressbar_inference_status_image = customtkinter.CTkLabel(self.image_frame, text="0%(0/0)", font=("Microsoft GothicNeo Bold", 12), anchor = "e")
        self.progressbar_inference_status_image.grid(row=2, column=2, padx=(300,0))

        self.progressbar_inference_info_image = customtkinter.CTkLabel(self.image_frame, text="FPS: -- \n Tiempo de inicio: --:--:-- \n Tiempo de inferencia: --:--:--", font=("Microsoft GothicNeo Bold", 12))
        self.progressbar_inference_info_image.grid(row=3, column=2, padx=(0,200))

        self.image_inference_reset_button = customtkinter.CTkButton(self.image_frame, text="Resetear", command = self.image_inference_reset_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.image_inference_reset_button.grid(row=3, column=2, padx=(200,0), pady=(0,0))

        self.image_inference_start_button = customtkinter.CTkButton(self.image_frame, text="Empezar", command = self.image_inference_start_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.image_inference_start_button.grid(row=4, column=2, padx=(0,200), pady=(0,0))

        self.image_inference_infographic_button = customtkinter.CTkButton(self.image_frame, text="Infografía", command = self.image_inference_infographic_button_event, anchor="center", height= 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.image_inference_infographic_button.grid(row=4, column=2, padx=(200,0), pady=(0,0))
        

        # create video frame
        self.video_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.video_frame.grid_columnconfigure(5, weight=1)
        self.video_frame.grid_rowconfigure(5, weight=1)

        self.source_selector_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_recursos_imagen.png")), size=(335, 53))
        self.source_selector = customtkinter.CTkLabel(self.video_frame, text="", image=self.source_selector_image)
        self.source_selector.grid(row=0, column=0, columnspan=2, padx=(0,20), pady=(19,0),sticky = "we")

        self.icon_upload_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_dark.png")), size=(45, 45))
        self.video_source_selector_button_1 = customtkinter.CTkButton(self.video_frame, text="SELECCIONAR CARPETA \n DE RECURSOS", image=self.icon_upload_files_image, compound="bottom", command= self.video_source_selector_button_1_event, height = 250)
        self.video_source_selector_button_1.grid(row=1, column=0, pady=(20,0)) 

        self.video_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Elementos seleccionados")
        self.video_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.video_select_source_folder_files = "Sin seleccionar" #Default value
        self.video_source_scrollable_elements = customtkinter.CTkLabel(master = self.video_source_scrollable_frame, text = self.video_select_source_folder_files, anchor="center")
        self.video_source_scrollable_elements.grid(row=0, column=0, padx = (55,0))

       
        self.source_save_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_guardado_imagen.png")), size=(335, 53))
        self.source_save = customtkinter.CTkLabel(self.video_frame, text="", image=self.source_save_image)
        self.source_save.grid(row=2, column=0, columnspan=2, padx=(0,20), pady=(20,0),sticky = "we")

        self.icon_download_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_dark.png")), size=(45, 45))
        self.video_source_save_button_1 = customtkinter.CTkButton(self.video_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.video_source_save_button_1_event, height=165)
        self.video_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.video_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.video_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.video_select_folder = "Sin seleccionar" #Default value
        self.video_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.video_source_save_scrollable_frame, text = self.video_select_folder, anchor = "w", compound="left")
        self.video_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))

       
        self.visual_inference_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "visualizacion_inferencia_imagen.png")), size=(442, 53))
        self.visual_inference = customtkinter.CTkLabel(self.video_frame, text="", image=self.visual_inference_image)
        self.visual_inference.grid(row=0, column=2, padx=(0,0), pady=(19,0),sticky = "we")
        
        self.video_inference_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "inferencia_image_light.png")), dark_image=Image.open(os.path.join(self.image_path, "inferencia_image_dark.png")),size=(439, 247))
        self.inference_video_image = customtkinter.CTkLabel(self.video_frame, text="", image=self.video_inference_image, compound="center", fg_color=('#3b8ed0','#3b8ed0')
                                                                                              , corner_radius=8, height=262)
        self.inference_video_image.grid(row=1, column=2, padx=(0,0), pady=(20, 0))

        self.progressbar_inference_video = customtkinter.CTkLabel(self.video_frame, text="PROGRESO:", font=("Microsoft GothicNeo Bold", 12), anchor = "w")
        self.progressbar_inference_video.grid(row=2, column=2, padx=(0,325))

        self.progressbar_video_frame = customtkinter.CTkProgressBar(self.video_frame, height = 20)
        self.progressbar_video_frame.grid(row=2, column=2, padx=(0,40)) #padx=(10, 20), pady=(10, 10), 
        self.progressbar_video_frame.set(0)

        self.progressbar_inference_status_video = customtkinter.CTkLabel(self.video_frame, text="0%(0/0)", font=("Microsoft GothicNeo Bold", 12), anchor = "e")
        self.progressbar_inference_status_video.grid(row=2, column=2, padx=(300,0))

        self.progressbar_inference_info_video = customtkinter.CTkLabel(self.video_frame, text="FPS: -- \n Tiempo de inicio: --:--:-- \n Tiempo de inferencia: --:--:--", font=("Microsoft GothicNeo Bold", 12))
        self.progressbar_inference_info_video.grid(row=3, column=2, padx=(0,200))

        self.video_inference_reset_button = customtkinter.CTkButton(self.video_frame, text="Resetear", command = self.video_inference_reset_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.video_inference_reset_button.grid(row=3, column=2, padx=(200,0), pady=(0,0))

        self.video_inference_start_button = customtkinter.CTkButton(self.video_frame, text="Empezar", command = self.video_inference_start_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.video_inference_start_button.grid(row=4, column=2, padx=(0,200), pady=(0,0))

        self.video_inference_infographic_button = customtkinter.CTkButton(self.video_frame, text="Infografía", command = self.video_inference_infographic_button_event, anchor="center", height= 50) #, image=self.icon_image_image, compound="bottom", command= self.video_source_save_button_1_event) # image=self.icon_image_image,
        self.video_inference_infographic_button.grid(row=4, column=2, padx=(200,0), pady=(0,0))
        

        # create streaming frame
        self.streaming_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.streaming_frame.grid_columnconfigure(5, weight=1)
        self.streaming_frame.grid_rowconfigure(5, weight=1)

        self.source_selector_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_recursos_imagen.png")), size=(335, 53))
        self.source_selector = customtkinter.CTkLabel(self.streaming_frame, text="", image=self.source_selector_image)
        self.source_selector.grid(row=0, column=0, columnspan=2, padx=(0,20), pady=(19,0),sticky = "we")

        self.icon_upload_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "subir_archivos_imagen_dark.png")), size=(45, 45))
        self.streaming_source_selector_button_1 = customtkinter.CTkButton(self.streaming_frame, text="INGRESAR LINK DEL EN \n VIVO", image=self.icon_upload_files_image, compound="bottom", command= self.streaming_source_selector_button_1_event, height = 165)
        self.streaming_source_selector_button_1.grid(row=1, column=0, pady=(20,0)) 

        self.streaming_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="Link ingresado", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.streaming_select_source_folder_files = "Sin ingresar" #Default value
        self.streaming_source_scrollable_elements = customtkinter.CTkLabel(master = self.streaming_source_scrollable_frame, text = self.streaming_select_source_folder_files, anchor="center")
        self.streaming_source_scrollable_elements.grid(row=0, column=0, padx = (75,0))

       
        self.source_save_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "seleccion_guardado_imagen.png")), size=(335, 53))
        self.source_save = customtkinter.CTkLabel(self.streaming_frame, text="", image=self.source_save_image)
        self.source_save.grid(row=2, column=0, columnspan=2, padx=(0,20), pady=(20,0),sticky = "we")

        self.icon_download_files_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_light.png")),
                                                 dark_image=Image.open(os.path.join(self.image_path, "bajar_archivos_imagen_dark.png")), size=(45, 45))
        self.streaming_source_save_button_1 = customtkinter.CTkButton(self.streaming_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.streaming_source_save_button_1_event, height=165)
        self.streaming_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.streaming_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.streaming_select_folder = "Sin seleccionar" #Default value
        self.streaming_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.streaming_source_save_scrollable_frame, text = self.streaming_select_folder, anchor = "w", compound="left")
        self.streaming_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))

       
        self.visual_inference_image = customtkinter.CTkImage(Image.open(os.path.join(self.image_path, "visualizacion_inferencia_imagen.png")), size=(442, 53))
        self.visual_inference = customtkinter.CTkLabel(self.streaming_frame, text="", image=self.visual_inference_image)
        self.visual_inference.grid(row=0, column=2, padx=(0,0), pady=(19,0),sticky = "we")
        
        self.streaming_inference_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(self.image_path, "inferencia_image_light.png")), dark_image=Image.open(os.path.join(self.image_path, "inferencia_image_dark.png")),size=(439, 247))
        self.inference_streaming_image = customtkinter.CTkLabel(self.streaming_frame, text="", image=self.streaming_inference_image, compound="center", fg_color=('#3b8ed0','#3b8ed0')
                                                                                              , corner_radius=8, height=262)
        self.inference_streaming_image.grid(row=1, column=2, padx=(0,0), pady=(20, 0))

        self.progressbar_inference_streaming = customtkinter.CTkLabel(self.streaming_frame, text="PROGRESO:", font=("Microsoft GothicNeo Bold", 12), anchor = "w")
        self.progressbar_inference_streaming.grid(row=2, column=2, padx=(0,325))

        self.progressbar_streaming_frame = customtkinter.CTkProgressBar(self.streaming_frame, height = 20)
        self.progressbar_streaming_frame.grid(row=2, column=2, padx=(0,40)) #padx=(10, 20), pady=(10, 10), 
        self.progressbar_streaming_frame.set(0)

        self.progressbar_inference_status_streaming = customtkinter.CTkLabel(self.streaming_frame, text="0%(0/0)", font=("Microsoft GothicNeo Bold", 12), anchor = "e")
        self.progressbar_inference_status_streaming.grid(row=2, column=2, padx=(300,0))

        self.progressbar_inference_info_streaming = customtkinter.CTkLabel(self.streaming_frame, text="FPS: -- \n Tiempo de inicio: --:--:-- \n Tiempo de inferencia: --:--:--", font=("Microsoft GothicNeo Bold", 12))
        self.progressbar_inference_info_streaming.grid(row=3, column=2, padx=(0,200))

        self.streaming_inference_reset_button = customtkinter.CTkButton(self.streaming_frame, text="Resetear", command = self.streaming_inference_reset_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.streaming_source_save_button_1_event) # image=self.icon_image_image,
        self.streaming_inference_reset_button.grid(row=3, column=2, padx=(200,0), pady=(0,0))

        self.streaming_inference_start_button = customtkinter.CTkButton(self.streaming_frame, text="Empezar", command = self.streaming_inference_start_button_event, anchor="center", height = 50) #, image=self.icon_image_image, compound="bottom", command= self.streaming_source_save_button_1_event) # image=self.icon_image_image,
        self.streaming_inference_start_button.grid(row=4, column=2, padx=(0,200), pady=(0,0))

        self.streaming_inference_infographic_button = customtkinter.CTkButton(self.streaming_frame, text="Infografía", command = self.streaming_inference_infographic_button_event, anchor="center", height= 50) #, image=self.icon_image_image, compound="bottom", command= self.streaming_source_save_button_1_event) # image=self.icon_image_image,
        self.streaming_inference_infographic_button.grid(row=4, column=2, padx=(200,0), pady=(0,0))

        # select default frame
        self.select_frame_by_name("home")
        self.state_1_streaming = 0
        self.state_2_streaming = 0
        self.state_1_video = 0
        self.state_2_video = 0
        self.state_1_image = 0
        self.state_2_image = 0
        self.start_infographic_video = 0
        self.streaming_select_source_folder = "Sin ingresar"
        self.streaming_inference_start_button.configure(state="disabled")
        self.streaming_inference_infographic_button.configure(state="disabled")
        self.video_select_source_folder = "Sin seleccionar"
        self.video_inference_start_button.configure(state="disabled")
        self.video_inference_infographic_button.configure(state="disabled")
        self.image_select_source_folder = "Sin seleccionar"
        self.image_inference_start_button.configure(state="disabled")
        self.image_inference_infographic_button.configure(state="disabled")
        


    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=('#0081c0', '#144870') if name == "home" else "transparent", border_width=0, corner_radius=8)
        self.image_button.configure(fg_color=('#0081c0', '#144870') if name == "image" else "transparent", border_width=0, corner_radius=8)
        self.video_button.configure(fg_color=('#0081c0', '#144870') if name == "video" else "transparent",border_width=0, corner_radius=8)
        self.streaming_button.configure(fg_color=('#0081c0', '#144870') if name == "streaming" else "transparent", border_width=0, corner_radius=8)

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "image":
            self.image_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.image_frame.grid_forget()
        if name == "video":
            self.video_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.video_frame.grid_forget()
        if name == "streaming":
            self.streaming_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.streaming_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def image_button_event(self):
        self.select_frame_by_name("image")

    def video_button_event(self):
        self.select_frame_by_name("video")

    def streaming_button_event(self):
        self.select_frame_by_name("streaming")
    
    def change_appearance_mode_event(self, new_appearance_mode: str):
        self.appearance = {"Claro":"Light", "Oscuro":"Dark","Sistema":"System"}
        customtkinter.set_appearance_mode(self.appearance[new_appearance_mode]) 

#Functions image
    def image_source_selector_button_1_event(self):
        if self.state_2_image == 1:
            self.state_1_image = 1
            self.image_inference_start_button.configure(state="enabled")
            self.image_inference_infographic_button.configure(state="disabled")       
        else:
            self.state_1_image = 1
        self.image_select_source_folder = filedialog.askdirectory()
        print(self.image_select_source_folder)
        self.image_select_source_folder_files = []
        self.image_select_source_folder_files.clear()
        for file in os.listdir(self.image_select_source_folder):
            if os.path.isfile(os.path.join(self.image_select_source_folder, file)):
                self.image_select_source_folder_files.append(file)
        print(self.image_select_source_folder_files)
        self.image_source_scrollable_frame.destroy()
        self.image_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Elementos seleccionados")#, orientation="horizontal")
        self.image_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        for i in range(len(self.image_select_source_folder_files)):
            self.image_source_scrollable_elements = customtkinter.CTkLabel(master = self.image_source_scrollable_frame, text = self.image_select_source_folder_files[i], anchor = "w", compound="left")
            self.image_source_scrollable_elements.grid(row=i, column=0, padx=(0,0))


    def image_source_save_button_1_event(self):
        self.image_select_folder = filedialog.askdirectory()
        print(self.image_select_folder)
        self.image_source_save_scrollable_frame.destroy()
        self.image_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.image_source_save_scrollable_frame.grid(row=3, column=1, rowspan=2, padx=(0,50), pady=(30,0))
        self.image_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.image_source_save_scrollable_frame, text = self.image_select_folder, anchor = "w", compound="left")
        self.image_source_save_scrollable_folder.grid(row=0, column=0, padx=(0,0))
        if self.state_1_image == 1:
            self.state_2_image = 1
            self.image_inference_start_button.configure(state="enabled")
            self.image_inference_infographic_button.configure(state="disabled")
        else:
            self.state_2_image = 1


    def image_inference_reset_button_event(self):
        self.image_source_scrollable_frame.destroy()
        self.image_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Elementos seleccionados")
        self.image_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.image_select_source_folder_files = "Sin seleccionar" #Default value
        self.image_source_scrollable_elements = customtkinter.CTkLabel(master = self.image_source_scrollable_frame, text = self.image_select_source_folder_files, anchor="center")
        self.image_source_scrollable_elements.grid(row=0, column=0, padx = (55,0))
        
        self.image_source_save_button_1 = customtkinter.CTkButton(self.image_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.image_source_save_button_1_event, height=165)
        self.image_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.image_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.image_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.image_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.image_select_folder = "Sin seleccionar" #Default value
        self.image_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.image_source_save_scrollable_frame, text = self.image_select_folder, anchor = "w", compound="left")
        self.image_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))

        self.state_2_image = 0
        self.state_1_image = 0
        self.image_inference_start_button.configure(state="disabled")
        self.image_inference_infographic_button.configure(state="disabled")

    
    def image_inference_start_button_event(self):
        self.image_inference_start_button.configure(state="disabled")
        self.image_inference_infographic_button.configure(state="enabled")
    
    
    def image_inference_infographic_button_event(self):
        self.image_inference_start_button.configure(state="enabled")
        self.image_inference_infographic_button.configure(state="disabled")        
    

#Functions video
    def video_source_selector_button_1_event(self):
        if self.state_2_video == 1:
            self.state_1_video = 1
            self.video_inference_start_button.configure(state="enabled")
            self.video_inference_infographic_button.configure(state="disabled")       
        else:
            self.state_1_video = 1
        self.video_select_source_folder = filedialog.askdirectory()
        print(self.video_select_source_folder)
        self.video_select_source_folder_files = []
        self.video_select_source_folder_files.clear()
        for file in os.listdir(self.video_select_source_folder):
            if os.path.isfile(os.path.join(self.video_select_source_folder, file)):
                self.video_select_source_folder_files.append(file)
        print(self.video_select_source_folder_files)
        self.video_source_scrollable_frame.destroy()
        self.video_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Elementos seleccionados")#, orientation="horizontal")
        self.video_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        for i in range(len(self.video_select_source_folder_files)):
            self.video_source_scrollable_elements = customtkinter.CTkLabel(master = self.video_source_scrollable_frame, text = self.video_select_source_folder_files[i], anchor = "w", compound="left")
            self.video_source_scrollable_elements.grid(row=i, column=0, padx=(0,0))
    

    def video_source_save_button_1_event(self):
        self.video_select_folder = filedialog.askdirectory()
        print(self.video_select_folder)
        self.video_source_save_scrollable_frame.destroy()
        self.video_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.video_source_save_scrollable_frame.grid(row=3, column=1, rowspan=2, padx=(0,50), pady=(30,0))
        self.video_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.video_source_save_scrollable_frame, text = self.video_select_folder, anchor = "w", compound="left")
        self.video_source_save_scrollable_folder.grid(row=0, column=0, padx=(0,0))
        if self.state_1_video == 1:
            self.state_2_video = 1
            self.video_inference_start_button.configure(state="enabled")
            self.video_inference_infographic_button.configure(state="disabled")
        else:
            self.state_2_video = 1

    
    def video_inference_reset_button_event(self):
        self.video_source_scrollable_frame.destroy()
        self.video_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Elementos seleccionados")
        self.video_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.video_select_source_folder_files = "Sin seleccionar" #Default value
        self.video_source_scrollable_elements = customtkinter.CTkLabel(master = self.video_source_scrollable_frame, text = self.video_select_source_folder_files, anchor="center")
        self.video_source_scrollable_elements.grid(row=0, column=0, padx = (55,0))
        
        self.video_source_save_button_1 = customtkinter.CTkButton(self.video_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.video_source_save_button_1_event, height=165)
        self.video_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.video_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.video_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.video_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.video_select_folder = "Sin seleccionar" #Default value
        self.video_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.video_source_save_scrollable_frame, text = self.video_select_folder, anchor = "w", compound="left")
        self.video_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))

        self.state_2_video = 0
        self.state_1_video = 0
        self.video_inference_start_button.configure(state="disabled")
        self.video_inference_infographic_button.configure(state="disabled")
        self.inference_video_image.configure(image=self.video_inference_image)
        self.progressbar_inference_info_video.configure(text="FPS: -- \n Tiempo de inicio: --:--:-- \n Tiempo de inferencia: --:--:--")



    def video_inference_start_button_event(self):
        self.video_inference_start_button.configure(state="disabled")
        self.video_inference_infographic_button.configure(state="enabled")
        app.update_idletasks()
        self.inference_func()
        

        	
    def video_inference_infographic_button_event(self):
        self.video_inference_start_button.configure(state="enabled")
        self.video_inference_infographic_button.configure(state="disabled")
        app.update_idletasks()
        self.infographic_func()
            

#Functions streaming
    def streaming_source_selector_button_1_event(self):
        if self.state_2_streaming == 1:
            self.state_1_streaming = 1
            self.streaming_inference_start_button.configure(state="enabled")
            self.streaming_inference_infographic_button.configure(state="disabled")       
        else:
            self.state_1_streaming = 1
        self.streaming_select_source_folder = customtkinter.CTkInputDialog(text="Ingresa el link del en vivo", title="Ingreso de link")
        self.streaming_select_source_folder = self.streaming_select_source_folder.get_input()
        print(self.streaming_select_source_folder)
        self.streaming_select_source_folder_files = []
        self.streaming_select_source_folder_files.clear()
        self.streaming_select_source_folder_files = self.streaming_select_source_folder
        self.streaming_source_scrollable_frame.destroy()
        self.streaming_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="En vivo seleccionado", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.streaming_source_scrollable_elements = customtkinter.CTkLabel(master = self.streaming_source_scrollable_frame, text = self.streaming_select_source_folder_files, anchor = "w", compound="left")
        self.streaming_source_scrollable_elements.grid(row=0, column=0, padx=(0,0))


    def streaming_source_save_button_1_event(self):
        self.streaming_select_folder = filedialog.askdirectory()
        print(self.streaming_select_folder)
        self.streaming_source_save_scrollable_frame.destroy()
        self.streaming_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_save_scrollable_frame.grid(row=3, column=1, rowspan=2, padx=(0,50), pady=(30,0))
        self.streaming_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.streaming_source_save_scrollable_frame, text = self.streaming_select_folder, anchor = "w", compound="left")
        self.streaming_source_save_scrollable_folder.grid(row=0, column=0, padx=(0,0))
        if self.state_1_streaming == 1:
            self.state_2_streaming = 1
            self.streaming_inference_start_button.configure(state="enabled")
            self.streaming_inference_infographic_button.configure(state="disabled")
        else:
            self.state_2_streaming = 1


    def streaming_inference_reset_button_event(self):
        self.streaming_source_scrollable_frame.destroy()
        self.streaming_source_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="En vivo seleccionado", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_scrollable_frame.grid(row=1, column=1, padx=(0,60), pady=(20,0))
        self.streaming_select_source_folder_files = "Sin ingresar" #Default value
        self.streaming_source_scrollable_elements = customtkinter.CTkLabel(master = self.streaming_source_scrollable_frame, text = self.streaming_select_source_folder_files, anchor="center")
        self.streaming_source_scrollable_elements.grid(row=0, column=0, padx = (75,0))
        
        self.streaming_source_save_button_1 = customtkinter.CTkButton(self.streaming_frame, text="SELECCIONAR CARPETA \n DE GUARDADO", image=self.icon_download_files_image, compound="bottom", command= self.streaming_source_save_button_1_event, height=165)
        self.streaming_source_save_button_1.grid(row=3, column=0, rowspan = 2, padx=20, pady=(30,0))
        self.streaming_source_save_scrollable_frame = customtkinter.CTkScrollableFrame( self.streaming_frame, label_text="Carpeta seleccionada", height = 100, orientation= "horizontal", width=222)
        self.streaming_source_save_scrollable_frame.grid(row=3, column=1, rowspan = 2, padx=(0,50), pady=(30,0))
        self.streaming_select_folder = "Sin seleccionar" #Default value
        self.streaming_source_save_scrollable_folder = customtkinter.CTkLabel(master = self.streaming_source_save_scrollable_frame, text = self.streaming_select_folder, anchor = "w", compound="left")
        self.streaming_source_save_scrollable_folder.grid(row=0, column=0, padx=(65,0))

        self.state_2_streaming = 0
        self.state_1_streaming = 0
        self.streaming_inference_start_button.configure(state="disabled")
        self.streaming_inference_infographic_button.configure(state="disabled")

    
    def streaming_inference_start_button_event(self):
        self.streaming_inference_start_button.configure(state="disabled")
        self.streaming_inference_infographic_button.configure(state="enabled")
    
    
    def streaming_inference_infographic_button_event(self):
        self.streaming_inference_start_button.configure(state="enabled")
        self.streaming_inference_infographic_button.configure(state="disabled") 

#Inference Function    
    @smart_inference_mode()
    def run(self):
        
	#Variables
        weights = str(ROOT / 'best.pt')  # model path or triton URL
        source= self.opt_1.source  # file/dir/URL/glob/screen/0(webcam)
        data=str(ROOT / 'data/coco128.yaml')  # dataset.yaml path
        imgsz=(640, 640)  # inference size (height, width)
        conf_thres=0.25  # confidence threshold
        iou_thres=0.45  # NMS IOU threshold
        max_det=1000  # maximum detections per image
        device=''  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        view_img=False  # show results
        save_txt=True  # save results to *.txt
        save_conf=False  # save confidences in --save-txt labels
        save_crop=False  # save cropped prediction boxes
        nosave=False  # do not save images/videos
        classes=None  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms=False  # class-agnostic NMS
        augment=False  # augmented inference
        visualize=False  # visualize features
        update=False  # update all models
        project= str(self.opt_1.project)  # save results to project/name
        name='exp'  # save results to project/name
        exist_ok=False  # existing project/name ok, do not increment
        line_thickness=3  # bounding box thickness (pixels)
        hide_labels=False  # hide labels
        hide_conf=False  # hide confidences
        half=False  # Use FP16 half-precision inference
        dnn=False  # use OpenCV DNN for ONNX inference
        vid_stride=1  # video frame-rate stride
        
        source = str(source)
        save_img = not nosave and not source.endswith('.txt')  # save inference images
        is_file = Path(source).suffix[1:] in (IMG_FORMATS + VID_FORMATS)
        is_url = source.lower().startswith(('rtsp://', 'rtmp://', 'http://', 'https://'))
        webcam = source.isnumeric() or source.endswith('.txt') or (is_url and not is_file)
        screenshot = source.lower().startswith('screen')
        if is_url and is_file:
            source = check_file(source)  # download

        # Directories
        print(project,end='\n==================================\n')
        save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
        (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

        # Load model
        device = select_device(device)
        model = DetectMultiBackend(weights=ROOT/'best.pt', device=device, dnn=dnn, data=data, fp16=half)
        stride, names, pt = model.stride, model.names, model.pt
        imgsz = check_img_size(imgsz, s=stride)  # check image size

        # Dataloader
        bs = 1  # batch_size
        if webcam:
            view_img = check_imshow(warn=True)
            dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
            bs = len(dataset)
        elif screenshot:
            #dataset = LoadScreenshots(**LSargs)
            dataset = LoadScreenshots(source, img_size=imgsz, stride=stride, auto=pt)
        else:
            dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
        vid_path, vid_writer = [None] * bs, [None] * bs

        # Run inference
        model.warmup(imgsz=(1 if pt or model.triton else bs, 3, *imgsz))  # warmup
        seen, windows, dt = 0, [], (Profile(), Profile(), Profile())
        
        # Active xlsx
        rowm = 1
        dect_1 = "Process dont pass"
        infer_1 = "Process dont pass"
        workbook = Workbook()
        sheet = workbook.active
        
        # Print inferenced frames per second (Only for video or stream inference)
        count_frames = 1
        infer_2 = 0
        inference_video_stream_active = 0   #0 for not, 1 active
        
        # Print start time
        start_time = datetime.datetime.now()
        LOGGER.info(f"Start time: {start_time}{''}")
        time_per_frames_inference = datetime.datetime.now()
        
            
        for path, im, im0s, vid_cap, s in dataset:

            app.update_idletasks()
            

        
            with dt[0]:
                im = torch.from_numpy(im).to(model.device)
                im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
                im /= 255  # 0 - 255 to 0.0 - 1.0
                if len(im.shape) == 3:
                    im = im[None]  # expand for batch dim

            # Inference
            with dt[1]:
                visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
                pred = model(im, augment=augment, visualize=visualize)

            # NMS
            with dt[2]:
                pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)

            # Second-stage classifier (optional)
            # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)

            # Process predictions
            for i, det in enumerate(pred):  # per image
                app.update_idletasks()
                seen += 1
                if webcam:  # batch_size >= 1
                    p, im0, frame = path[i], im0s[i].copy(), dataset.count
                    s += f'{i}: '
                else:
                    p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)

                p = Path(p)  # to Path
                save_path = str(save_dir / p.name)  # im.jpg
                txt_path = str(save_dir / 'labels' / p.stem) + ('' if dataset.mode == 'image' else f'_{frame}')  # im.txt
                s += '%gx%g ' % im.shape[2:]  # print string
                s_m = ''
                gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
                imc = im0.copy() if save_crop else im0  # for save_crop
                annotator = Annotator(im0, line_width=line_thickness, example=str(names))
                if len(det):
                    # Rescale boxes from img_size to im0 size
                    det[:, :4] = scale_boxes(im.shape[2:], det[:, :4], im0.shape).round()

                    # Print results
                    for c in det[:, 5].unique():
                        app.update_idletasks()
                        n = (det[:, 5] == c).sum()  # detections per class
                        s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string
                        s_m += f"{n} {names[int(c)]}{'s' * (n > 1)}, " # concatenated string with detections per class
                        dect_1 = s_m	# variable of detections is updated

                    # Write results
                    for *xyxy, conf, cls in reversed(det):
                        app.update_idletasks()
                        if save_txt:  # Write to file
                            xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                            line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                            with open(f'{txt_path}.txt', 'a') as f:
                                f.write(('%g ' * len(line)).rstrip() % line + '\n')

                        if save_img or save_crop or view_img:  # Add bbox to image
                            c = int(cls)  # integer class
                            label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                            annotator.box_label(xyxy, label, color=colors(c, True))
                        if save_crop:
                            save_one_box(xyxy, imc, file=save_dir / 'crops' / names[c] / f'{p.stem}.jpg', BGR=True)

                # Stream results
                im0 = annotator.result()
                if view_img:
                    if platform.system() == 'Linux' and p not in windows:
                        windows.append(p)
                        cv2.namedWindow(str(p), cv2.WINDOW_NORMAL | cv2.WINDOW_KEEPRATIO)  # allow window resize (Linux)
                        cv2.resizeWindow(str(p), im0.shape[1], im0.shape[0])
                    cv2.imshow(str(p), im0)
                    cv2.waitKey(1)  # 1 millisecond

                # Save results (image with detections)
                if save_img:
                    if dataset.mode == 'image':
                        cv2.imwrite(save_path, im0)
                    else:  # 'video' or 'stream'
                        inference_video_stream_active = 1
                        if vid_path[i] != save_path:  # new video
                            vid_path[i] = save_path
                            if isinstance(vid_writer[i], cv2.VideoWriter):
                                vid_writer[i].release()  # release previous video writer
                            if vid_cap:  # video
                                fps = vid_cap.get(cv2.CAP_PROP_FPS)
                                w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                                h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                            else:  # stream
                                fps, w, h = 30, im0.shape[1], im0.shape[0]
                            save_path = str(Path(save_path).with_suffix('.mp4'))  # force *.mp4 suffix on results videos
                            vid_writer[i] = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (w, h))
                        vid_writer[i].write(im0)
                        
            
            # Print time (inference-only)
            time_per_inference1 = time_per_frames_inference
            time_per_inference2 = datetime.datetime.now()
            time_per_frames_inference_seconds = (time_per_inference2 - time_per_inference1).total_seconds()
            
            # Write in cells of xlxs file (detections and inference time)
            if not dt[1].dt:
                infer_1 = "No detections"
            else:
                infer_1 = time_per_frames_inference_seconds
                infer_2 = infer_2 + infer_1
                if infer_2 >= 1: #At least one second of inference has passed
                    frames_per_second = count_frames/infer_2
                    inference_minutes, inference_seconds = divmod(infer_2,60)
                    inference_hours, inference_minutes = divmod(inference_minutes,60)
                    LOGGER.info(f"{s}{'' if len(det) else '(no detections), '}Time inference per this frame: {time_per_frames_inference_seconds* 1E0:.3f}s, Total inference time: {infer_2 * 1E0:.3f}{''}s, Inferred frames per second: {frames_per_second * 1E0:.3f}{''}")
                    self.progressbar_inference_info_video.configure(text=('FPS: ' + str(round(frames_per_second,3))) + '\n' + 'Tiempo de inicio: ' + str(start_time.time())[:15 -7] + '\n' + 'Tiempo de inferencia: ' + str(int(inference_hours)) + ':' + str(int(inference_minutes)) + ':' + str(int(inference_seconds)))
                    
                    cv2.imwrite('im0.jpg',im0)
                    cv2.waitKey(0)
                    im0_path = str(ROOT)
                    im0_img = customtkinter.CTkImage(Image.open(os.path.join(im0_path, "im0.jpg")), size=(442, 247))
                    self.inference_video_image.configure(image=im0_img)
                    for palabra in s.split():
                        if '(' in palabra and ')' in palabra:
                            inicio = palabra.index('(')+1
                            fin = palabra.index(')')
                            contenido = palabra[inicio:fin]
                            if '/' in contenido:
                                frame_number_1, frame_number_2 = contenido.split('/')
                    frame_numbers=int(frame_number_1)/int(frame_number_2)
                    self.progressbar_video_frame.set(frame_numbers)
                    self.progressbar_inference_status_video.configure(text=str(int(frame_numbers*100))+ '%' + '' + '(' + str(frame_number_1) + '/' + str(frame_number_2) + ')')        
                    app.update_idletasks()
                    if os.path.isfile('im0.jpg'):
                        os.remove('im0.jpg')
                else:   
                        LOGGER.info(f"{s}{'' if len(det) else '(no detections), '}{time_per_frames_inference_seconds* 1E0:.3f}s, inference time: {infer_2* 1E0:.3f}{''}s")
                count_frames+=1
            if not len(det):
                dect_1 = "No detections"
                infer_1 = "No detections"
            sheet.cell(row=rowm, column=1).value = dect_1
            sheet.cell(row=rowm, column=2).value = infer_1
            rowm+=1
            clasdect_nam = "classesdetect.xlsx"
            clasdect_dic = str(save_dir / clasdect_nam)
            workbook.save(filename=clasdect_dic)
            
            time_per_frames_inference = datetime.datetime.now()
            
            
            
        # Print results
        t = tuple(x.t / seen * 1E3 for x in dt)  # speeds per image
        LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {(1, 3, *imgsz)}' % t)
        final_time = datetime.datetime.now()
        LOGGER.info(f"Final time: {final_time}{''}")
        delta_time = (final_time - start_time).total_seconds()
        running_minutes, running_seconds = divmod(delta_time,60)
        running_hours, running_minutes = divmod(running_minutes,60)
        if inference_video_stream_active == 1:
            #frames_per_second = math.trunc(1000/t[1])
            parts = str(frames_per_second).split(".")
            third_decimals = parts[1][:3]
            frames_per_second_third_decimals = parts[0] + "." + third_decimals
            LOGGER.info(f"Inferred frames per second: {float(frames_per_second_third_decimals)}{'' if frames_per_second>0 else '(no detections)'}, Total time of inference: {int(running_hours)}{''} hours, {int(running_minutes)}{''} minutes and {int(running_seconds)}{''} seconds")
            sheet.cell(row=1, column=4).value = "Inferred frames per second: " + str(frames_per_second_third_decimals)
            clasdect_nam = "classesdetect.xlsx"
            clasdect_dic = str(save_dir / clasdect_nam)
            workbook.save(filename=clasdect_dic)
        else:
            LOGGER.info(f"Total time of inference: {int(running_hours)}{''} hours, {int(running_minutes)}{''} minutes and {int(running_seconds)}{''} seconds")
        sheet.cell(row=1, column=3).value = start_time
        sheet.cell(row=2, column=3).value = final_time
        running_hours, running_minutes, running_seconds = str(int(running_hours)), str(int(running_minutes)), str(int(running_seconds))
        sheet.cell(row=3, column=3).value = running_hours + " hours " + running_minutes + " minutes and " + running_seconds + " seconds" 
        clasdect_nam = "classesdetect.xlsx"
        clasdect_dic = str(save_dir / clasdect_nam)
        workbook.save(filename=clasdect_dic)

        if save_txt or save_img:
            s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
            LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
        if update:
            strip_optimizer(weights[0])  # update model (to fix SourceChangeWarning)

    def parse_opt(self):
        parser = argparse.ArgumentParser()
        parser.add_argument('--weights', nargs='+', type=str, default=ROOT/'best.pt', help='model path or triton URL') #
        parser.add_argument('--source', type=str, default=ROOT / 'data/images', help='file/dir/URL/glob/screen/0(webcam)')
        parser.add_argument('--data', type=str, default=ROOT / 'data/coco128.yaml', help='(optional) dataset.yaml path')
        parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default=[640], help='inference size h,w')
        parser.add_argument('--conf-thres', type=float, default=0.25, help='confidence threshold')
        parser.add_argument('--iou-thres', type=float, default=0.45, help='NMS IoU threshold')
        parser.add_argument('--max-det', type=int, default=1000, help='maximum detections per image')
        parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
        parser.add_argument('--view-img', action='store_true', help='show results') # default='True',
        parser.add_argument('--save-txt', default='True',action='store_true', help='save results to *.txt')
        parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
        parser.add_argument('--save-crop', action='store_true', help='save cropped prediction boxes')
        parser.add_argument('--nosave', action='store_true', help='do not save images/videos')
        parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --classes 0, or --classes 0 2 3')
        parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
        parser.add_argument('--augment', action='store_true', help='augmented inference')
        parser.add_argument('--visualize', action='store_true', help='visualize features')
        parser.add_argument('--update', action='store_true', help='update all models')
        parser.add_argument('--project', default=ROOT / 'runs/detect', help='save results to project/name')
        parser.add_argument('--name', default='exp', help='save results to project/name')
        parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
        parser.add_argument('--line-thickness', default=3, type=int, help='bounding box thickness (pixels)')
        parser.add_argument('--hide-labels', default=False, action='store_true', help='hide labels')
        parser.add_argument('--hide-conf', default=False, action='store_true', help='hide confidences')
        parser.add_argument('--half', action='store_true', help='use FP16 half-precision inference')
        parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
        parser.add_argument('--vid-stride', type=int, default=1, help='video frame-rate stride')
        self.opt = parser.parse_args()
        self.opt.imgsz *= 2 if len(self.opt.imgsz) == 1 else 1  # expand
        self.opt.source = self.video_select_source_folder
        self.opt.project = self.video_select_folder
        print_args(vars(self.opt))
        return self.opt
           

    def inference_func(self):
        self.opt_1 = self.parse_opt()
        check_requirements(exclude=('tensorboard', 'thop'))
        #print(opt)
        self.run()

    def infographic_func(self):
        dir = self.video_select_folder + '/exp' + '/labels/'
        rootdir=os.listdir(dir)
        rootdir.sort(key=lambda s: int(s.split('.txt')[0][-1]))

        newdf=pd.DataFrame()

        filas={0:'awning-tricycle',
            1:'biclycle',
            2:'bus',
            3:'car',
            4:'motor',
            5:'pedestrian',
            6:'people',
            7:'tricycle',
            8:'truck',
            9:'van'}

        for i in range(len(rootdir)):
		df = pd.read_csv(dir+rootdir[i], sep=" ",header=None)
		r=df[0].value_counts()
		newdf=pd.concat([newdf,r],axis=1, ignore_index=True)
        newdf=newdf.sort_index()
        newdf=newdf.rename(index=filas)

        newdf

        newdf1=newdf.mean(axis=1)

        import numpy as np

        newdf=pd.concat([newdf,newdf1.apply(np.ceil)],axis=1,ignore_index=True)
        newdf

        import matplotlib.pyplot as plt

        plt.barh(newdf.index,newdf.loc[:,newdf.columns.values[-1]])

        newdf2=newdf.sum(axis=1)-newdf[newdf.columns.values[-1]]
        newdf2

        sumas=pd.DataFrame(newdf2)
        sumas

        h=sumas.index.to_numpy().tolist()
        h

        f=sumas[0].to_numpy().tolist()
        f

        fig =plt.figure(figsize = (4,4))
        ax11 = fig.add_subplot(111)
        w,l,p = ax11.pie(f, autopct='%1.1f%%', pctdistance=1, radius = 0.5)

        pctdists = [.1,.25,.4,.55,.7,.85,1,1.15,1.3]

        for t,d in zip(p, pctdists):
            xi,yi = t.get_position()
            ri = np.sqrt(xi**2+yi**2)
            phi = np.arctan2(yi,xi)
            x = d*ri*np.cos(phi)
            y = d*ri*np.sin(phi)
            t.set_position((x,y))
        ax11.legend(labels=h, loc="upper center", ncol=2)
        plt.axis('equal')
        plt.savefig(str(self.video_select_folder)+'/exp/'+'infographic.jpg')
        infographic_image = cv2.imread(str(self.video_select_folder)+'/exp/'+'infographic.jpg')
        cv2.imshow('Infografía',infographic_image)
        cv2.waitKey(0)

if __name__ == "__main__":
    app = App()
    app.mainloop()
